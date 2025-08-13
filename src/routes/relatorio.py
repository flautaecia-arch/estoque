from flask import Blueprint, send_file, jsonify
from src.models.user import db
from src.models.produto import Produto
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os
from datetime import datetime
from sqlalchemy import func

relatorio_bp = Blueprint('relatorio', __name__)

class RelatorioPDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 16)
        self.cell(0, 10, 'Relatório de Estoque', 0, 1, 'C')
        self.set_font('Arial', '', 10)
        self.cell(0, 10, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 1, 'C')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')

@relatorio_bp.route('/relatorio/pdf', methods=['GET'])
def gerar_relatorio_pdf():
    try:
        # Buscar todos os produtos ordenados por código
        produtos = Produto.query.order_by(Produto.codigo.asc()).all()
        
        if not produtos:
            return jsonify({'error': 'Nenhum produto encontrado para gerar relatório'}), 404

        # Criar PDF
        pdf = RelatorioPDF()
        pdf.add_page()
        
        # Título do relatório
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Relatório de Estoque - Detalhado por Lote', 0, 1, 'L')
        pdf.ln(5)
        
        # Cabeçalho da tabela
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(25, 8, 'Código', 1, 0, 'C')
        pdf.cell(60, 8, 'Nome do Produto', 1, 0, 'C')
        pdf.cell(25, 8, 'Lote', 1, 0, 'C')
        pdf.cell(25, 8, 'Validade', 1, 0, 'C')
        pdf.cell(20, 8, 'Qtd', 1, 0, 'C')
        pdf.cell(25, 8, 'Cadastro', 1, 1, 'C')
        
        # Dados dos produtos agrupados por código
        pdf.set_font('Arial', '', 8)
        codigo_atual = None
        subtotal_codigo = 0
        total_geral = 0
        
        for produto in produtos:
            # Verificar se mudou o código do produto
            if codigo_atual != produto.codigo:
                # Se não é o primeiro produto, mostrar subtotal do código anterior
                if codigo_atual is not None:
                    pdf.set_font('Arial', 'B', 8)
                    pdf.cell(135, 6, f'Subtotal {codigo_atual}:', 1, 0, 'R')
                    pdf.cell(20, 6, str(subtotal_codigo), 1, 0, 'C')
                    pdf.cell(25, 6, '', 1, 1, 'C')
                    pdf.ln(2)
                    pdf.set_font('Arial', '', 8)
                
                codigo_atual = produto.codigo
                subtotal_codigo = 0
            
            # Verificar se precisa de nova página
            if pdf.get_y() > 250:
                pdf.add_page()
                # Repetir cabeçalho
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(25, 8, 'Código', 1, 0, 'C')
                pdf.cell(60, 8, 'Nome do Produto', 1, 0, 'C')
                pdf.cell(25, 8, 'Lote', 1, 0, 'C')
                pdf.cell(25, 8, 'Validade', 1, 0, 'C')
                pdf.cell(20, 8, 'Qtd', 1, 0, 'C')
                pdf.cell(25, 8, 'Cadastro', 1, 1, 'C')
                pdf.set_font('Arial', '', 8)
            
            # Adicionar linha do produto
            pdf.cell(25, 6, str(produto.codigo), 1, 0, 'C')
            pdf.cell(60, 6, produto.nome[:25], 1, 0, 'L')
            pdf.cell(25, 6, produto.lote, 1, 0, 'C')
            pdf.cell(25, 6, produto.validade.strftime('%d/%m/%Y'), 1, 0, 'C')
            pdf.cell(20, 6, str(produto.quantidade), 1, 0, 'C')
            pdf.cell(25, 6, produto.data_cadastro.strftime('%d/%m/%Y'), 1, 1, 'C')
            
            subtotal_codigo += produto.quantidade
            total_geral += produto.quantidade
        
        # Mostrar subtotal do último código
        if codigo_atual is not None:
            pdf.set_font('Arial', 'B', 8)
            pdf.cell(135, 6, f'Subtotal {codigo_atual}:', 1, 0, 'R')
            pdf.cell(20, 6, str(subtotal_codigo), 1, 0, 'C')
            pdf.cell(25, 6, '', 1, 1, 'C')
        
        # Linha de separação
        pdf.ln(5)
        
        # Total geral
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(135, 8, 'TOTAL GERAL:', 1, 0, 'R')
        pdf.cell(20, 8, str(total_geral), 1, 0, 'C')
        pdf.cell(25, 8, '', 1, 1, 'C')
        
        # Salvar PDF em memória
        pdf_output = io.BytesIO()
        pdf_string = pdf.output(dest='S')
        if isinstance(pdf_string, str):
            pdf_string = pdf_string.encode('latin-1')
        pdf_output.write(pdf_string)
        pdf_output.seek(0)
        
        return send_file(
            pdf_output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro ao gerar relatório PDF: {str(e)}'}), 500

@relatorio_bp.route('/relatorio/excel', methods=['GET'])
def gerar_relatorio_excel():
    try:
        # Buscar todos os produtos ordenados por código
        produtos = Produto.query.order_by(Produto.codigo.asc()).all()
        
        # Buscar resumo por código
        resumo = db.session.query(
            Produto.codigo,
            Produto.nome,
            func.sum(Produto.quantidade).label('quantidade_total'),
            func.count(Produto.id).label('total_lotes')
        ).group_by(Produto.codigo, Produto.nome).order_by(Produto.codigo.asc()).all()

        # Criar workbook
        wb = Workbook()
        
        # Aba de resumo
        ws_resumo = wb.active
        ws_resumo.title = "Resumo por Código"
        
        # Cabeçalho do resumo
        headers_resumo = ['Código', 'Nome do Produto', 'Quantidade Total', 'Total de Lotes']
        for col, header in enumerate(headers_resumo, 1):
            cell = ws_resumo.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados do resumo
        for row, item in enumerate(resumo, 2):
            ws_resumo.cell(row=row, column=1, value=item.codigo)
            ws_resumo.cell(row=row, column=2, value=item.nome)
            ws_resumo.cell(row=row, column=3, value=item.quantidade_total)
            ws_resumo.cell(row=row, column=4, value=item.total_lotes)
        
        # Ajustar largura das colunas
        ws_resumo.column_dimensions['A'].width = 15
        ws_resumo.column_dimensions['B'].width = 40
        ws_resumo.column_dimensions['C'].width = 18
        ws_resumo.column_dimensions['D'].width = 15
        
        # Aba de detalhes (agora com subtotais e total geral)
        ws_detalhes = wb.create_sheet(title="Detalhes por Lote")
        
        # Título do relatório
        ws_detalhes.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        title_cell = ws_detalhes.cell(row=1, column=1, value="Relatório de Estoque - Detalhado por Lote")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Cabeçalho da tabela
        headers_detalhes = ["Código", "Nome do Produto", "Lote", "Validade", "Qtd", "Cadastro"]
        for col, header in enumerate(headers_detalhes, 1):
            cell = ws_detalhes.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados dos produtos agrupados por código
        row_idx = 4
        codigo_atual = None
        subtotal_codigo = 0
        total_geral = 0
        
        for produto in produtos:
            # Verificar se mudou o código do produto
            if codigo_atual != produto.codigo:
                # Se não é o primeiro produto, mostrar subtotal do código anterior
                if codigo_atual is not None:
                    ws_detalhes.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                    ws_detalhes.cell(row=row_idx, column=1, value=f"Subtotal {codigo_atual}:").font = Font(bold=True)
                    ws_detalhes.cell(row=row_idx, column=5, value=subtotal_codigo).font = Font(bold=True)
                    row_idx += 1
                    ws_detalhes.row_dimensions[row_idx].height = 5 # Espaçamento
                    row_idx += 1
                
                codigo_atual = produto.codigo
                subtotal_codigo = 0
            
            # Adicionar linha do produto
            ws_detalhes.cell(row=row_idx, column=1, value=produto.codigo)
            ws_detalhes.cell(row=row_idx, column=2, value=produto.nome)
            ws_detalhes.cell(row=row_idx, column=3, value=produto.lote)
            ws_detalhes.cell(row=row_idx, column=4, value=produto.validade.strftime("%d/%m/%Y"))
            ws_detalhes.cell(row=row_idx, column=5, value=produto.quantidade)
            ws_detalhes.cell(row=row_idx, column=6, value=produto.data_cadastro.strftime("%d/%m/%Y %H:%M"))
            row_idx += 1
            
            subtotal_codigo += produto.quantidade
            total_geral += produto.quantidade
        
        # Mostrar subtotal do último código
        if codigo_atual is not None:
            ws_detalhes.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            ws_detalhes.cell(row=row_idx, column=1, value=f"Subtotal {codigo_atual}:").font = Font(bold=True)
            ws_detalhes.cell(row=row_idx, column=5, value=subtotal_codigo).font = Font(bold=True)
            row_idx += 1
        
        # Linha de separação
        row_idx += 1
        ws_detalhes.row_dimensions[row_idx].height = 10 # Espaçamento
        row_idx += 1
        
        # Total geral
        ws_detalhes.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        total_cell = ws_detalhes.cell(row=row_idx, column=1, value="TOTAL GERAL:")
        total_cell.font = Font(bold=True, size=12)
        total_cell.alignment = Alignment(horizontal="right")
        ws_detalhes.cell(row=row_idx, column=5, value=total_geral).font = Font(bold=True, size=12)
        
        # Ajustar largura das colunas
        ws_detalhes.column_dimensions["A"].width = 15
        ws_detalhes.column_dimensions["B"].width = 40
        ws_detalhes.column_dimensions["C"].width = 15
        ws_detalhes.column_dimensions["D"].width = 12
        ws_detalhes.column_dimensions["E"].width = 12
        ws_detalhes.column_dimensions["F"].width = 18
        
        # Salvar Excel em memória
        excel_output = io.BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)
        
        return send_file(
            excel_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro ao gerar relatório Excel: {str(e)}'}), 500

